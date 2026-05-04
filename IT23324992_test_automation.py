"""
run_tests_fixed.py
Runs the Playwright automation against the Singlish chat translator,
reading from IT23324992_updated.xlsx with explicit column indices.

Column layout (1-indexed):
  1 = Test Case ID
  2 = Input length type
  3 = Input                  <-- read from here
  4 = Expected output        <-- compare against
  5 = Actual Output          <-- write here
  6 = Status                 <-- write FAIL/PASS here
  7 = Singlish input types covered
"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='backslashreplace')

import time, re, os
from pathlib import Path
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.cell.cell import MergedCell

# ── CONFIG ──────────────────────────────────────────────────────────────
EXCEL_PATH   = Path(__file__).resolve().parent / "IT23324992.xlsx"
SHEET_NAME   = " Test cases"
URL          = "https://www.pixelssuite.com/chat-translator"

# Explicit column indices (1-based) – never auto-detected
COL_INPUT    = 3
COL_EXPECTED = 4
COL_ACTUAL   = 5
COL_STATUS   = 6

HEADER_ROW   = 1      # row that contains column headers
DATA_START   = 2      # first data row

WAIT_MS        = 7000   # ms to wait after clicking Transliterate
RETRIES        = 10     # how many times to poll for output
RETRY_WAIT_MS  = 1500
TYPE_DELAY_MS  = 25
SAVE_EVERY     = 5      # save Excel every N rows
# ────────────────────────────────────────────────────────────────────────


def log(msg):
    print(msg, flush=True)


def dismiss_overlays(page):
    for role, name in [
        ("button", re.compile(r"^(Accept|I Agree|Agree|OK|Got it|Accept all)$", re.I)),
    ]:
        try:
            btn = page.get_by_role(role, name=name).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(400)
        except Exception:
            pass


def clear_input(page, loc):
    for _ in range(3):
        try: loc.click(timeout=2000)
        except Exception: pass
        try: page.keyboard.press("Control+A"); page.keyboard.press("Backspace")
        except Exception: pass
        try: loc.fill("")
        except Exception: pass
        try:
            if loc.input_value() == "": return
        except Exception: pass
        try:
            loc.evaluate("(el) => { el.value=''; el.dispatchEvent(new Event('input',{bubbles:true})); }")
            if loc.input_value() == "": return
        except Exception: pass
        page.wait_for_timeout(200)


def type_input(page, loc, text):
    clear_input(page, loc)
    if TYPE_DELAY_MS > 0:
        loc.click(timeout=2000)
        loc.type(text, delay=TYPE_DELAY_MS)
    else:
        loc.fill(text)


def read_output(loc):
    for fn in [
        lambda: loc.input_value(),
        lambda: loc.inner_text(),
        lambda: loc.text_content(),
        lambda: loc.evaluate("(el) => 'value' in el ? el.value : el.innerText || ''"),
    ]:
        try:
            v = fn()
            if v and str(v).strip():
                return str(v).strip()
        except Exception:
            pass
    return ""


def find_locators(page, timeout_ms=60000):
    """Find input textarea, output textarea and Transliterate button."""
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        dismiss_overlays(page)
        try:
            inp = page.locator('textarea[placeholder*="English"]').first
            out = page.locator('textarea[placeholder*="Sinhala"]').first
            if inp.count() > 0 and out.count() > 0 and inp.is_visible() and out.is_visible():
                btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.I)).first
                return inp, out, btn
        except Exception:
            pass
        try:
            textareas = [page.locator("textarea").nth(i)
                         for i in range(page.locator("textarea").count())
                         if page.locator("textarea").nth(i).is_visible()]
            if len(textareas) >= 2:
                btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.I)).first
                return textareas[0], textareas[1], btn
        except Exception:
            pass
        page.wait_for_timeout(500)
    raise RuntimeError("Could not find chat UI textareas after timeout.")


def get_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell


def set_cell(ws, row, col, value):
    get_cell(ws, row, col).value = value


def main():
    log(f"Loading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    log(f"Sheet: '{ws.title}' | Rows: {ws.max_row} | Input col: {COL_INPUT} | Expected col: {COL_EXPECTED}")

    # Verify headers at fixed columns
    for idx, name in [(COL_INPUT, "Input"), (COL_EXPECTED, "Expected output"),
                      (COL_ACTUAL, "Actual Output"), (COL_STATUS, "Status")]:
        actual_header = ws.cell(row=HEADER_ROW, column=idx).value
        log(f"  Col {idx} header = [{actual_header}] (expected: [{name}])")

    total_rows = ws.max_row - DATA_START + 1
    log(f"\nStarting {total_rows} test cases...\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=0)
        page = browser.new_page()
        page.set_default_timeout(60000)

        log(f"Opening: {URL}")
        page.goto(URL, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=30000)
        except Exception:
            pass

        try:
            inp_loc, out_loc, btn_loc = find_locators(page, timeout_ms=60000)
            log("UI locators found OK.\n")
        except Exception as e:
            log(f"ERROR: {e}")
            browser.close()
            return

        processed = 0
        for row in range(DATA_START, ws.max_row + 1):
            tc_id = get_cell(ws, row, 1).value
            singlish = get_cell(ws, row, COL_INPUT).value
            expected = get_cell(ws, row, COL_EXPECTED).value
            actual_current = get_cell(ws, row, COL_ACTUAL).value

            if not singlish:
                continue

            if actual_current and actual_current not in ["", "(empty)", "පරිවර්තනය දෝෂ සහිතයි", "UI Error", "None"]:
                continue

            singlish = str(singlish).strip()
            expected_str = str(expected).strip() if expected else ""

            log(f"[{row-1:02d}/50] [{tc_id}] Input: {singlish[:60]}")

            try:
                dismiss_overlays(page)
                type_input(page, inp_loc, singlish)

                try:
                    btn_loc.click(timeout=5000)
                except Exception:
                    log("  (Transliterate button not found, trying Enter)")
                    inp_loc.press("Enter")

                page.wait_for_timeout(WAIT_MS)

                actual = read_output(out_loc)

                set_cell(ws, row, COL_ACTUAL, actual)

                if expected_str:
                    status = "PASS" if actual.strip() == expected_str.strip() else "FAIL"
                else:
                    status = "COLLECTED"

                set_cell(ws, row, COL_STATUS, status)
                log(f"  -> Actual: {actual[:60]} | Status: {status}")

                processed += 1
                if SAVE_EVERY > 0 and processed % SAVE_EVERY == 0:
                    wb.save(str(EXCEL_PATH))
                    log(f"  [Saved at row {row}]")

            except Exception as e:
                log(f"  ERROR: {e}")
                set_cell(ws, row, COL_STATUS, "UI Error")
                set_cell(ws, row, COL_ACTUAL, "")

        browser.close()

    wb.save(str(EXCEL_PATH))
    log(f"\nAll done! {processed} rows processed. Saved to: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
